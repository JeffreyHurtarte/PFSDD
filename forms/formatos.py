from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.page import PageMargins
from django.conf import settings
from .models import Tipo, Rol, Usuario, Area, Etapa, Ambiente, Proyecto, Formulario, Detalle_Formulario, Bitacora, Workspace, Programa, Region_Cics, Tipo_Programa, Version_Cobol, Tipo_Formulario
import os

def form_traslado_fases():
    file_path = os.path.join(settings.MEDIA_ROOT, 'traslado_fases.xlsx')
    if os.path.exists(file_path):
        os.remove(file_path)

    wb = Workbook()
    ws = wb.active

    ws.oddHeader.left.text = "Plataforma ZSeries"
    ws.oddHeader.left.size = 10
    ws.oddHeader.left.font = "Courier New"
    ws.oddHeader.left.bold = True

    ws.oddHeader.center.text = "TRASLADO DE FASE NUEVO"
    ws.oddHeader.center.size = 14
    ws.oddHeader.center.font = "Courier New"
    ws.oddHeader.center.bold = True

    ws.oddHeader.right.text = "FORMA: MF-03\nVersión 01-2021"
    ws.oddHeader.right.size = 10
    ws.oddHeader.right.font = "Courier New"
    ws.oddHeader.right.bold = True

    fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    for row in ws.iter_rows(min_row=2, max_row=49, min_col=1, max_col=21):
        for cell in row:
            cell.fill = fill
            if cell.row == 2:
                cell.border = Border(top=Side(style='thin'), left=Side(style='thin') if cell.column == 1 else None, right=Side(style='thin') if cell.column == 21 else None)
            elif cell.row == 49:
                cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin') if cell.column == 1 else None, right=Side(style='thin') if cell.column == 21 else None)
            elif cell.column == 1:
                cell.border = Border(left=Side(style='thin'))
            elif cell.column == 21:
                cell.border = Border(right=Side(style='thin'))

    column_widths = {
        'A': 14, 'B': 42, 'C': 48, 'D': 44, 'E': 44, 'F': 37,
        'G': 42, 'H': 48, 'I': 42, 'J': 42, 'K': 33, 'L': 42,
        'M': 48, 'N': 17, 'O': 42, 'P': 33, 'Q': 42, 'R': 42,
        'S': 17, 'T': 37, 'U': 9  
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width / 7.5

    row_heights = {
        1:   6, 2:   4, 3:   2, 4:  26, 5:  16, 6:  16, 
        7:  24, 8:  24, 9:  24, 10: 21, 11: 21, 12:  9, 
        13: 32, 14: 20, 15: 20, 16: 20, 17: 20, 18: 20, 
        19: 20, 20: 20, 21: 20, 22: 20, 23: 20, 24: 20, 
        25: 20, 26: 20, 27: 20, 28: 20, 29: 20, 30: 20, 
        31: 20, 32: 20, 33: 20, 34: 20, 35: 20, 36: 20, 
        37:  1, 38: 20, 39: 67, 40: 10, 41: 38, 42:  2, 
        43: 21, 44: 18, 45:  1, 46: 21, 47:  8, 48: 21, 
        49:  9, 50:  9
    }

    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height * 0.75

    ws.page_margins = PageMargins(
        left=0.6, right=0.6, top=1.9, bottom=2.3, header=0.8, footer=1.5
    )

    ws.merge_cells('B4:D4')
    cell = ws['B4']
    cell.value = "RATIONAL:"
    cell.font = Font(name='Courier New', size=10, bold=True, italic=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('B5:D6')
    cell = ws['B5']
    cell.value = "PROYECTO:"
    cell.font = Font(name='Courier New', size=10, bold=True, italic=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    for row in [7, 8, 9]:
        ws.merge_cells(f'B{row}:D{row}')
        cell = ws[f'B{row}']
        cell.value = "LIB./WORKSPACE:"
        cell.font = Font(name='Courier New', size=10, bold=True, italic=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')

    proyecto = Proyecto.objects.last()
    
    ws.merge_cells('E4:H4')
    cell = ws['E4']
    cell.value = proyecto.rational if proyecto else ""
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(5, 9):
        ws.cell(row=4, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('B10:E10')
    cell = ws['B10']
    cell.value = "FECHA DE SOLICITUD:"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('L10:P10')
    cell = ws['L10']
    cell.value = "FECHA DE TRASLADO:"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('B11:E11')
    cell = ws['B11']
    cell.value = "HORA DE SOLICITUD:"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('L11:P11')
    cell = ws['L11']
    cell.value = "HORA DE TRASLADO:"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('E5:S6')
    cell = ws['E5']
    cell.value = proyecto.nombre if proyecto else ""
    for row in range(5, 7):
        for col in range(5, 20):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    workspace = Workspace.objects.last()

    for row in [7, 8, 9]:
        ws.merge_cells(f'E{row}:S{row}')
        cell = ws['E7']
        cell.value = workspace.nombre if workspace else ""
        for col in range(5, 20):
            cell = ws.cell(row=row, column=col)
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('F10:J10')
    cell = ws['F10']
    cell.value = proyecto.fecha if proyecto else ""
    for col in range(6, 11):
        cell = ws.cell(row=10, column=col)
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('Q10:T10')
    for col in range(17, 21):
        cell = ws.cell(row=10, column=col)
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('F11:J11')
    for col in range(6, 11):
        cell = ws.cell(row=11, column=col)
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('Q11:T11')
    for col in range(17, 21):
        cell = ws.cell(row=11, column=col)
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('B13:C13')
    cell = ws['B13']
    cell.value = "PROGRAMA A TRASLADAR"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(2, 4):
        ws.cell(row=13, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['D13'].value = "COBOL 4.2"
    ws['D13'].font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['D13'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['E13'].value = "COBOL 6.3"
    ws['E13'].font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['E13'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws['F13'].value = "NCU1"
    ws['F13'].font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['F13'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('G13:H13')
    cell = ws['G13']
    cell.value = "REGION DE CICS"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(7, 9):
        ws.cell(row=13, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('I13:J13')
    cell = ws['I13']
    cell.value = "TIPO DE PROGRAMA"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(9, 11):
        ws.cell(row=13, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('K13:L13')
    cell = ws['K13']
    cell.value = "MAPSET / COPYS"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(11, 13):
        ws.cell(row=13, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('M13:T13')
    cell = ws['M13']
    cell.value = "Traslado WS Sistemas:"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    programa = Programa.objects.last()
    region = Region_Cics.objects.last()

    for row in range(14, 37):
        ws.merge_cells(f'B{row}:C{row}')
        cell = ws['B14']
        cell.value = programa.nombre if programa else ""
        cell = ws[f'B{row}']
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(2, 4):
            ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws[f'D{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'D{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws[f'E{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell = ws['E14']
        cell.value = "X"
        ws[f'E{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws[f'F{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'F{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells(f'G{row}:H{row}')
        cell = ws['G14']
        cell.value = region.nombre if region else ""
        cell = ws[f'G{row}']
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(7, 9):
            ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells(f'I{row}:J{row}')
        cell = ws[f'I{row}']
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(9, 11):
            ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells(f'K{row}:L{row}')
        cell = ws[f'K{row}']
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(11, 13):
            ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('M14:Q14')
    cell = ws['M14']
    cell.value = "Fecha traslado WS:"
    cell.font = Font(name='Courier New', size=10)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('M15:Q15')
    cell = ws['M15']
    cell.value = "Hora traslado WS:"
    cell.font = Font(name='Courier New', size=10)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('M17:O18')
    cell = ws['M17']
    cell.value = "Nombre y firma:"
    cell.font = Font(name='Courier New', size=10)
    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    ws.merge_cells('M19:P19')
    cell = ws['M19']
    cell.value = "Antigüedad:"
    cell.font = Font(name='Courier New', size=10)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('M20:P20')
    cell = ws['M20']
    cell.value = "Corporativo:"
    cell.font = Font(name='Courier New', size=10)
    cell.alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('R14:T14')
    for col in range(18, 21):
        cell = ws.cell(row=14, column=col)
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('R15:T15')
    for col in range(18, 21):
        cell = ws.cell(row=15, column=col)
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('Q19:T19')
    for col in range(17, 21):
        cell = ws.cell(row=19, column=col)
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('P17:T18')
    cell = ws['P17']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in range(17, 19):
        for col in range(16, 21):
            ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('Q20:T20')
    for col in range(17, 21):
        cell = ws.cell(row=20, column=col)
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in range(22, 35):
        for col in range(14, 20):
            cell = ws.cell(row=row, column=col)
            border = Border()
            if row == 22:
                border = Border(top=Side(style='thin'))
            if row == 34:
                border = Border(bottom=Side(style='thin'))
            if col == 14:
                border = Border(left=Side(style='thin'))
            if col == 19:
                border = Border(right=Side(style='thin'))
            if row == 22 and col == 14:
                border = Border(top=Side(style='thin'), left=Side(style='thin'))
            if row == 22 and col == 19:
                border = Border(top=Side(style='thin'), right=Side(style='thin'))
            if row == 34 and col == 14:
                border = Border(bottom=Side(style='thin'), left=Side(style='thin'))
            if row == 34 and col == 19:
                border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
            cell.border = border

    ws.merge_cells('N22:S23')
    cell = ws['N22']
    cell.value = "USO EXCLUSIVO DEL AREA DE SISTEMAS"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in range(22, 24):
        for col in range(14, 20):
            cell = ws.cell(row=row, column=col)
            border = Border()
            if row == 22:
                border = Border(top=Side(style='thin'))
            if row == 23:
                border = Border(bottom=None)
            if col == 14:
                border = Border(left=Side(style='thin'))
            if col == 19:
                border = Border(right=Side(style='thin'))
            if row == 22 and col == 14:
                border = Border(top=Side(style='thin'), left=Side(style='thin'))
            if row == 22 and col == 19:
                border = Border(top=Side(style='thin'), right=Side(style='thin'))
            if row == 23 and col == 14:
                border = Border(left=Side(style='thin'))
            if row == 23 and col == 19:
                border = Border(right=Side(style='thin'))
            cell.border = border

    for row in range(25, 33):
        cell = ws[f'O{row}']
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('O33:P33')
    cell = ws['O33']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(15, 17):
        ws.cell(row=33, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    cell = ws['P25']
    cell.value = "Back up fase"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    cell = ws['P26']
    cell.value = "FTP fase"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    cell = ws['P27']
    cell.value = "Back up fuente"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    cell = ws['P28']
    cell.value = "FTP fuente"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    cell = ws['P29']
    cell.value = "Edición fuente"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    cell = ws['P30']
    cell.value = "Copia fuente"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    cell = ws['P31']
    cell.value = "Elimina fuente"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    cell = ws['P32']
    cell.value = "Workspace"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    cell = ws['Q33']
    cell.value = "Línea Base"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('N35:T35')
    cell = ws['N35']
    cell.value = "OBSERVACIONES:"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('N36:T36')
    cell = ws['N36']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(14, 21):
        ws.cell(row=36, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('B38:T38')
    cell = ws['B38']
    cell.value = "Descripción de programa y razón de traslado:"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('B39:T39')
    cell = ws['B39']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(2, 21):
        ws.cell(row=39, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('B41:E41')
    cell = ws['B41']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(2, 6):
        ws.cell(row=41, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('G41:J41')
    cell = ws['G41']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(7, 11):
        ws.cell(row=41, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('L41:O41')
    cell = ws['L41']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(12, 16):
        ws.cell(row=41, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('Q41:T41')
    cell = ws['Q41']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(17, 21):
        ws.cell(row=41, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('B43:E43')
    cell = ws['B43']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    ws.merge_cells('G43:J43')
    cell = ws['G43']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    ws.merge_cells('L43:O43')
    cell = ws['L43']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    ws.merge_cells('Q43:T43')
    cell = ws['Q43']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    ws.merge_cells('B44:E44')
    cell = ws['B44']
    cell.value = "Nombre de programador"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('G44:J44')
    cell = ws['G44']
    cell.value = "Vo. Bo. Documentación"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('L44:O44')
    cell = ws['L44']
    cell.value = "Vo. Bo. Coordinador"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('Q44:T44')
    cell = ws['Q44']
    cell.value = "Vo. Bo. Sistemas"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('B46:C46')
    cell = ws['B46']
    cell.value = "Antigüedad"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('G46:H46')
    cell = ws['G46']
    cell.value = "Antigüedad"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('L46:M46')
    cell = ws['L46']
    cell.value = "Antigüedad"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('Q46:R46')
    cell = ws['Q46']
    cell.value = "Antigüedad"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('D46:E46')
    cell = ws['D46']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(4, 6):
        ws.cell(row=46, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('I46:J46')
    cell = ws['I46']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(9, 11):
        ws.cell(row=46, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('N46:O46')
    cell = ws['N46']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(14, 16):
        ws.cell(row=46, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('S46:T46')
    cell = ws['S46']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(19, 21):
        ws.cell(row=46, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('B48:C48')
    cell = ws['B48']
    cell.value = "Corporativo"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('G48:H48')
    cell = ws['G48']
    cell.value = "Corporativo"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('L48:M48')
    cell = ws['L48']
    cell.value = "Corporativo"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('Q48:R48')
    cell = ws['Q48']
    cell.value = "Corporativo"
    cell.font = Font(name='Courier New', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('D48:E48')
    cell = ws['D48']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(4, 6):
        ws.cell(row=48, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('I48:J48')
    cell = ws['I48']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(9, 11):
        ws.cell(row=48, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('N48:O48')
    cell = ws['N48']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(14, 16):
        ws.cell(row=48, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('S48:T48')
    cell = ws['S48']
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(19, 21):
        ws.cell(row=48, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    wb.save(file_path)
    return file_path



def form_solicitud_programas():
    file_path = os.path.join(settings.MEDIA_ROOT, 'solicitud_programas.xlsx')
    if os.path.exists(file_path):
        os.remove(file_path)

    wb = Workbook()
    ws = wb.active

    ws.oddHeader.center.text = "SOLICITUD NOMBRE PROGRAMAS"
    ws.oddHeader.center.size = 14
    ws.oddHeader.center.font = "Courier New"
    ws.oddHeader.center.bold = True

    ws.oddHeader.right.text = "DESARROLLO ZSERIES\nFORMA: ÚNICA"
    ws.oddHeader.right.size = 10
    ws.oddHeader.right.font = "Courier New"
    ws.oddHeader.right.bold = True

    fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')

    for row in ws.iter_rows(min_row=1, max_row=37, min_col=1, max_col=14):
        for cell in row:
            cell.fill = fill
            if cell.row == 1:
                cell.border = Border(top=Side(style='thin'), left=Side(style='thin') if cell.column == 1 else None, right=Side(style='thin') if cell.column == 14 else None)
            elif cell.row == 37:
                cell.border = Border(bottom=Side(style='thin'), left=Side(style='thin') if cell.column == 1 else None, right=Side(style='thin') if cell.column == 14  else None)
            elif cell.column == 1:
                cell.border = Border(left=Side(style='thin'))
            elif cell.column == 14:
                cell.border = Border(right=Side(style='thin'))

    column_widths = {
        'A': 38, 'B':  5, 'C': 50, 'D': 54, 'E': 60, 'F': 100,
        'G': 65, 'H': 55, 'I': 65, 'J': 57, 'K': 32, 'L':  77,
        'M': 71, 'N': 37,  
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width / 7.5

    row_heights = {
        1:  30, 2:  15, 3:  18, 4:  24, 5:  24, 6:  93, 
        7:  25, 8:  25, 9:  25, 10: 25, 11: 25, 12: 25, 
        13: 25, 14: 25, 15: 25, 16: 25, 17: 25, 18: 25, 
        19: 25, 25: 25, 21: 25, 22: 25, 23: 25, 24: 25, 
        25: 25, 26: 25, 27: 25, 28: 25, 29: 25, 30: 25, 
        31: 25, 32: 25, 33: 25, 34: 18, 35: 18, 36: 18, 
        37: 18
    }

    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height * 0.75

    ws.page_margins = PageMargins(
        left=0.2362, right=0.2362, top=0.748, bottom=0.5906, header=0.3149, footer=0.5118
    )

    ws.merge_cells('C5:E5')
    cell = ws['C5']
    cell.value = "NOMBRE PROGRAMADOR:"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(3, 6):
        ws.cell(row=5, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('F5:M5')
    cell = ws['F5']
    cell.value = ""
    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for col in range(6, 14):
        ws.cell(row=5, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('C6:D6')
    cell = ws['C6']
    cell.value = "APLICACIÓN"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)    
    for col in range(3, 5):
        ws.cell(row=6, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    cell = ws['E6']
    cell.value = "PROGRAMADOR"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['E6'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
 
    cell = ws['F6']
    cell.value = "FRECUENCIA (D,M,O,U,S)"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['F6'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    cell = ws['G6']
    cell.value = "VERSIÓN"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['G6'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    cell = ws['H6']
    cell.value = "TIPO"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['H6'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    cell = ws['I6']
    cell.value = "CALIDAD"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['I6'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('J6:K6')
    cell = ws['J6']
    cell.value = "CANTIDAD"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)   
    for col in range(10, 12):
        ws.cell(row=6, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) 

    cell = ws['L6']
    cell.value = "MAPAS"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['L6'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))     

    cell = ws['M6']
    cell.value = "TRANSACCIÓN"
    cell.font = Font(name='Courier New', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)    
    ws['M6'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in range(7, 34):
        ws.merge_cells(f'C{row}:D{row}')
        cell = ws[f'C{row}']
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(3, 5):
            ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws[f'E{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'E{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws[f'F{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'F{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws[f'G{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'G{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws[f'H{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'H{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws[f'I{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'I{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells(f'J{row}:K{row}')
        cell = ws[f'J{row}']
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col in range(10, 12):
            ws.cell(row=row, column=col).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws[f'L{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'L{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'L{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws[f'M{row}'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        ws[f'M{row}'].border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[f'M{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb.save(file_path)
    return file_path


